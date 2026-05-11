from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "outputs" / "supplier-dataset"
PDF_PATH = OUTPUT_DIR / "bd_textile_exporters.pdf"
PDF_URL = "https://www.bangladeshtradeportal.gov.bd/kcfinder/upload/files/Stat_1639908873.pdf"
SYNTHETIC_XLSX = OUTPUT_DIR / "sourcery_synthetic_suppliers_500_additional.xlsx"
REAL_XLSX = OUTPUT_DIR / "sourcery_real_suppliers_200_public.xlsx"
COMBINED_JSON = OUTPUT_DIR / "sourcery_supplier_expansion_700.json"
COMBINED_CSV = OUTPUT_DIR / "sourcery_supplier_expansion_700.csv"

SYNTHETIC_BATCH = "buildfest-500-additional"
REAL_BATCH = "public-bd-exporters-200"

DB_COLUMNS = [
    "id",
    "name",
    "country",
    "city",
    "region",
    "category",
    "products",
    "description",
    "moq",
    "lead_time_days",
    "monthly_capacity",
    "unit_price_usd",
    "rating",
    "risk_level",
    "risk_score",
    "risk_notes",
    "bgmea_certified",
    "certifications",
    "payment_terms",
    "contact_name",
    "email",
    "phone",
    "website",
    "notes",
    "metadata",
]

EXPORT_COLUMNS = [
    "supplier_id",
    "name",
    "country",
    "city",
    "region",
    "category",
    "subcategory",
    "products",
    "description",
    "unit_price_usd",
    "moq",
    "lead_time_days",
    "monthly_capacity",
    "on_time_rate",
    "quality_rating",
    "risk_score",
    "risk_level",
    "bgmea_certified",
    "certifications",
    "payment_terms",
    "contact_name",
    "email",
    "phone",
    "website",
    "source_type",
    "source_url",
    "verified_at",
    "embedding_status",
    "metadata_json",
    "notes",
]


def load_env_file() -> dict[str, str]:
    env: dict[str, str] = {}
    path = ROOT / ".env.local"
    if not path.exists():
        return env

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def uuid_from_seed(seed: str) -> str:
    hex_value = hashlib.sha256(seed.encode("utf-8")).hexdigest()[:32]
    return f"{hex_value[:8]}-{hex_value[8:12]}-4{hex_value[13:16]}-a{hex_value[17:20]}-{hex_value[20:32]}"


def normalized_fraction(seed: str) -> float:
    return int(hashlib.sha256(seed.encode("utf-8")).hexdigest()[:8], 16) / 0xFFFFFFFF


def numeric(seed: str, low: float, high: float, decimals: int = 0) -> float:
    value = low + normalized_fraction(seed) * (high - low)
    return round(value, decimals)


def choose(values: list[str], index: int) -> str:
    return values[index % len(values)]


def slugify(text: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")
    return slug[:60] or "supplier"


def smart_title(text: str) -> str:
    words = text.split()
    titled: list[str] = []
    preserve = {"ltd", "pvt", "co", "bd", "usa", "uk", "llc", "inc", "hk"}
    for word in words:
        clean = re.sub(r"[^a-zA-Z0-9/&().-]", "", word)
        if not clean:
            continue
        if clean.lower() in preserve:
            titled.append(clean.upper() if len(clean) <= 3 else clean.title())
        elif clean.isupper() and len(clean) > 1:
            titled.append(clean.title())
        else:
            titled.append(clean)
    return " ".join(titled)


def strip_address_tail(text: str) -> str:
    patterns = [
        r"\b(?:Plot|House|Road|Sector|Block|Building|Tower|Plaza|Chamber|Bhaban|Complex|Center|Centre|Floor|Lane|Avenue|Industrial Area|DOHS|C/A|Bazar|Bagh|Uttara|Banani|Gulshan|Dhaka|Gazipur|Narayanganj|Chittagong|Chattogram|Bangladesh)\b",
        r"\b\d{1,4}[/-]?\d*\b",
    ]
    stripped = text
    for pattern in patterns:
        match = re.search(pattern, stripped, flags=re.I)
        if match and match.start() > 6:
            stripped = stripped[: match.start()].strip(" ,.-")
            break
    return stripped


COUNTRY_PLANS = [
    {
        "country": "Bangladesh",
        "region": "South Asia",
        "count": 80,
        "cities": ["Dhaka", "Gazipur", "Narayanganj", "Chattogram", "Khulna", "Cumilla", "Savar", "Sylhet"],
        "stems": ["Padma", "Jamuna", "Meghna", "Bengal", "Dhaka", "Sundarban", "Rupsha", "Shapla"],
        "certs": ["BGMEA", "BSCI", "OEKO-TEX", "WRAP", "BKMEA", "ISO 9001"],
        "ports": ["Chittagong", "Mongla"],
        "incoterms": ["FOB", "CFR"],
    },
    {
        "country": "India",
        "region": "South Asia",
        "count": 70,
        "cities": ["Tiruppur", "Mumbai", "Delhi", "Jaipur", "Bengaluru", "Kolkata", "Noida"],
        "stems": ["Deccan", "Jaipur", "Tiruppur", "Mumbai", "Indus", "Lotus", "Kolkata"],
        "certs": ["GOTS", "BSCI", "ISO 9001", "ISO 22716", "SA8000"],
        "ports": ["Nhava Sheva", "Chennai", "Mundra"],
        "incoterms": ["FOB", "EXW"],
    },
    {
        "country": "Pakistan",
        "region": "South Asia",
        "count": 40,
        "cities": ["Karachi", "Lahore", "Faisalabad", "Sialkot", "Multan"],
        "stems": ["Indus", "Karachi", "Punjab", "Sialkot", "Lahore"],
        "certs": ["BSCI", "OEKO-TEX", "ISO 9001", "Leather Working Group"],
        "ports": ["Karachi", "Port Qasim"],
        "incoterms": ["FOB", "CFR"],
    },
    {
        "country": "Vietnam",
        "region": "Southeast Asia",
        "count": 50,
        "cities": ["Ho Chi Minh City", "Hanoi", "Da Nang", "Hai Phong", "Binh Duong"],
        "stems": ["Saigon", "Mekong", "Lotus", "Hanoi", "Da Nang"],
        "certs": ["BSCI", "WRAP", "ISO 9001", "GRS"],
        "ports": ["Hai Phong", "Cat Lai"],
        "incoterms": ["FOB", "EXW"],
    },
    {
        "country": "China",
        "region": "East Asia",
        "count": 80,
        "cities": ["Shenzhen", "Guangzhou", "Ningbo", "Dongguan", "Hangzhou", "Foshan"],
        "stems": ["Pearl", "Shenzhen", "Guangzhou", "Ningbo", "Dragon", "Foshan"],
        "certs": ["ISO 9001", "CE", "RoHS", "BSCI", "FSC"],
        "ports": ["Shenzhen", "Guangzhou", "Ningbo"],
        "incoterms": ["FOB", "EXW"],
    },
    {
        "country": "Turkey",
        "region": "MENA",
        "count": 35,
        "cities": ["Istanbul", "Bursa", "Izmir", "Gaziantep"],
        "stems": ["Anatolia", "Istanbul", "Bursa", "Aegean", "Bosphorus"],
        "certs": ["OEKO-TEX", "GOTS", "ISO 9001", "Leather Working Group"],
        "ports": ["Istanbul", "Izmir"],
        "incoterms": ["FOB", "EXW"],
    },
    {
        "country": "Portugal",
        "region": "Europe",
        "count": 25,
        "cities": ["Porto", "Lisbon", "Braga", "Aveiro"],
        "stems": ["Porto", "Atlantic", "Braga", "Lusitania"],
        "certs": ["GOTS", "OEKO-TEX", "BSCI", "REACH"],
        "ports": ["Leixoes", "Lisbon"],
        "incoterms": ["EXW", "DAP"],
    },
    {
        "country": "Morocco",
        "region": "MENA",
        "count": 25,
        "cities": ["Casablanca", "Tangier", "Marrakesh", "Agadir"],
        "stems": ["Atlas", "Casablanca", "Tangier", "Sahara"],
        "certs": ["Ecocert", "ISO 22716", "BSCI", "ISO 9001"],
        "ports": ["Casablanca", "Tangier Med"],
        "incoterms": ["FOB", "EXW"],
    },
    {
        "country": "Indonesia",
        "region": "Southeast Asia",
        "count": 35,
        "cities": ["Jakarta", "Bandung", "Surabaya", "Semarang", "Denpasar"],
        "stems": ["Java", "Nusantara", "Bandung", "Bali", "Jakarta"],
        "certs": ["BSCI", "Rainforest Alliance", "ISO 9001", "GRS"],
        "ports": ["Jakarta", "Surabaya"],
        "incoterms": ["FOB", "EXW"],
    },
    {
        "country": "Egypt",
        "region": "MENA",
        "count": 20,
        "cities": ["Cairo", "Alexandria", "10th of Ramadan City", "Giza"],
        "stems": ["Nile", "Cairo", "Alex", "Pharaoh"],
        "certs": ["OEKO-TEX", "BSCI", "ISO 9001", "ISO 22000"],
        "ports": ["Alexandria", "Port Said"],
        "incoterms": ["FOB", "CFR"],
    },
    {
        "country": "Brazil",
        "region": "South America",
        "count": 20,
        "cities": ["Sao Paulo", "Curitiba", "Porto Alegre", "Campinas"],
        "stems": ["Sao Paulo", "Amazon", "Rio", "Curitiba"],
        "certs": ["GMP", "ISO 22716", "FSC", "BSCI"],
        "ports": ["Santos", "Paranagua"],
        "incoterms": ["FOB", "EXW"],
    },
    {
        "country": "Mexico",
        "region": "North America",
        "count": 20,
        "cities": ["Monterrey", "Guadalajara", "Tijuana", "Puebla"],
        "stems": ["Monterrey", "Aztec", "Puebla", "Baja"],
        "certs": ["ISO 9001", "BSCI", "HACCP", "FSC"],
        "ports": ["Veracruz", "Lazaro Cardenas"],
        "incoterms": ["EXW", "DAP"],
    },
]

CATEGORY_PRODUCTS = {
    "apparel": ["organic cotton hoodies", "denim jackets", "polo shirts", "activewear sets", "woven shirts", "modest fashion tops"],
    "beauty": ["face serum", "lip balm", "soap bars", "hair care kits", "argan skincare", "private label cosmetics"],
    "home": ["bedding", "towels", "ceramic tableware", "home decor", "rugs", "curtains"],
    "food": ["rice packs", "tea bags", "spice blends", "honey jars", "snack pouches", "coffee beans"],
    "accessories": ["jute tote bags", "canvas backpacks", "leather wallets", "gift bags", "phone cases", "travel pouches"],
    "packaging": ["shipping cartons", "paper bags", "cosmetic boxes", "labels", "jars", "airless bottles"],
    "electronics": ["chargers", "bluetooth modules", "sensor boards", "power banks", "cable kits"],
    "textiles": ["cotton fabric", "linen blends", "jute fabric rolls", "twill", "yarn"],
    "footwear": ["canvas sneakers", "leather sandals", "casual shoes", "kids footwear"],
    "industrial": ["metal brackets", "retail fixtures", "custom metal parts", "machine housings"],
}

CATEGORY_PRICE = {
    "apparel": (3.2, 12.8),
    "beauty": (1.0, 7.6),
    "home": (2.2, 18.5),
    "food": (0.9, 6.2),
    "accessories": (0.8, 9.4),
    "packaging": (0.12, 2.8),
    "electronics": (1.1, 11.0),
    "textiles": (0.55, 4.8),
    "footwear": (5.8, 22.0),
    "industrial": (3.5, 16.5),
}

CATEGORY_MOQ = {
    "apparel": (250, 2500, 50),
    "beauty": (500, 5000, 100),
    "home": (200, 1800, 50),
    "food": (400, 3000, 100),
    "accessories": (150, 1800, 50),
    "packaging": (1000, 10000, 100),
    "electronics": (500, 5000, 100),
    "textiles": (500, 4000, 100),
    "footwear": (300, 2000, 50),
    "industrial": (250, 3000, 50),
}

CATEGORY_SET = list(CATEGORY_PRODUCTS.keys())

BANGLADESH_CITIES = {
    "Dhaka": "Dhaka",
    "Gazipur": "Gazipur",
    "Narayangonj": "Narayanganj",
    "Narayanganj": "Narayanganj",
    "Savar": "Savar",
    "Ashulia": "Ashulia",
    "Mymenshingh": "Mymensingh",
    "Mymensingh": "Mymensingh",
    "Chittagong": "Chattogram",
    "Chattogram": "Chattogram",
    "Cumilla": "Cumilla",
    "Comilla": "Cumilla",
    "Sylhet": "Sylhet",
    "Fatullah": "Fatullah",
    "Mirpur": "Dhaka",
    "Uttara": "Dhaka",
    "Tejgaon": "Dhaka",
    "Banani": "Dhaka",
    "Gulshan": "Dhaka",
    "Badda": "Dhaka",
    "Dakhin Khan": "Dhaka",
    "DEPZ": "Savar",
}

PRODUCT_TERMS = {
    "apparel": [
        "t-shirt",
        "t shirt",
        "polo shirt",
        "shirt",
        "blouse",
        "jacket",
        "trouser",
        "pant",
        "dress",
        "pyjama",
        "sweater",
        "cardigan",
        "pullover",
        "legging",
        "hooded jacket",
        "tank top",
        "denim",
    ],
    "textiles": ["fabric", "yarn", "textile", "knitting", "dyeing", "finishing", "denim fabric"],
    "accessories": ["cap", "hats", "bag", "bags", "wallet", "belt"],
    "home": ["towel", "bathrobe", "sheet", "bedding", "curtain"],
    "footwear": ["shoe", "shoes", "sneaker", "sneakers", "sandals"],
}


def build_synthetic_suppliers(target_count: int) -> list[dict[str, Any]]:
    suppliers: list[dict[str, Any]] = []
    for plan in COUNTRY_PLANS:
        for index in range(plan["count"]):
            category = CATEGORY_SET[index % len(CATEGORY_SET)]
            products = [
                choose(CATEGORY_PRODUCTS[category], index),
                choose(CATEGORY_PRODUCTS[category], index + 1),
                choose(CATEGORY_PRODUCTS[category], index + 2),
            ]
            seed = f"{SYNTHETIC_BATCH}:{plan['country']}:{category}:{index}"
            price_low, price_high = CATEGORY_PRICE[category]
            moq_low, moq_high, moq_step = CATEGORY_MOQ[category]
            is_bd = plan["country"] == "Bangladesh"
            moq_raw = numeric(f"{seed}:moq", moq_low, moq_high)
            moq = int(round(moq_raw / moq_step) * moq_step)
            lead = int(numeric(f"{seed}:lead", 16 if is_bd else 22, 42 if is_bd else 58))
            capacity = int(numeric(f"{seed}:capacity", 18000, 220000))
            price = numeric(f"{seed}:price", price_low, price_high, 2)
            rating = numeric(f"{seed}:rating", 3.8, 4.9, 2)
            risk_score = int(numeric(f"{seed}:risk", 14 if is_bd else 18, 46 if is_bd else 62))
            risk_level = "low" if risk_score <= 30 else "medium" if risk_score <= 60 else "high"
            certs = list({choose(plan["certs"], index), choose(plan["certs"], index + 2), "ISO 9001"})
            stem = choose(plan["stems"], index)
            city = choose(plan["cities"], index)
            name = f"{stem} {city} {category.title()} Works {index + 1:03d}"
            description = (
                f"{city}, {plan['country']} supplier for {', '.join(products[:3])} with export documentation, quote-ready operations, "
                f"sample coordination, and SME-friendly production planning."
            )
            metadata = {
                "source_kind": "synthetic",
                "synthetic_batch": SYNTHETIC_BATCH,
                "sample_days": int(numeric(f"{seed}:sample", 4, 12)),
                "incoterms": plan["incoterms"],
                "port": choose(plan["ports"], index),
            }
            supplier = {
                "id": uuid_from_seed(seed),
                "name": name,
                "country": plan["country"],
                "city": city,
                "region": plan["region"],
                "category": category,
                "products": products,
                "description": description,
                "moq": moq,
                "lead_time_days": lead,
                "monthly_capacity": capacity,
                "unit_price_usd": price,
                "rating": rating,
                "risk_level": risk_level,
                "risk_score": risk_score,
                "risk_notes": f"Synthetic benchmark profile. Risk score {risk_score}/100 estimated from lead time, MOQ, and product complexity.",
                "bgmea_certified": bool(is_bd and category in {"apparel", "accessories", "textiles"} and index % 2 == 0),
                "certifications": certs,
                "payment_terms": choose(
                    [
                        "30% advance, 70% before shipment",
                        "LC at sight",
                        "50% deposit, 50% before dispatch",
                        "Net 30 for approved buyers",
                    ],
                    index,
                ),
                "contact_name": f"{choose(['Amina', 'Farhan', 'Nadia', 'Mahmud', 'Sara', 'Tariq', 'Priya', 'Minh'], index)} {choose(['Rahman', 'Karim', 'Islam', 'Hossain', 'Ahmed', 'Tran', 'Nair'], index + 1)}",
                "email": f"sales+{slugify(name)}@sourcery-demo.example",
                "phone": f"+880-{1000 + index:04d}-{2000 + index:04d}" if is_bd else None,
                "website": f"https://{slugify(name)}.example",
                "notes": f"Synthetic supplier row generated for Sourcery BuildFest demo coverage. Batch: {SYNTHETIC_BATCH}.",
                "metadata": metadata,
            }
            suppliers.append(supplier)
    return suppliers[:target_count]


def download_pdf() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if PDF_PATH.exists():
        return
    response = requests.get(PDF_URL, timeout=120)
    response.raise_for_status()
    PDF_PATH.write_bytes(response.content)


def clean_line(text: str) -> str:
    text = text.replace("\x00", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip(" -\u2022\t")


def extract_entries_from_pdf() -> list[list[str]]:
    download_pdf()
    reader = PdfReader(str(PDF_PATH))
    entries: list[list[str]] = []
    current: list[str] = []

    for page_index in range(110, min(len(reader.pages), 320)):
        page_text = reader.pages[page_index].extract_text() or ""
        lines = [clean_line(line) for line in page_text.splitlines()]
        for line in lines:
            if not line:
                continue
            start_match = re.match(r"^(\d{3,4})\s*(.+)$", line)
            if start_match and 100 <= int(start_match.group(1)) <= 3500 and re.search(r"[A-Za-z]", start_match.group(2)):
                if current:
                    entries.append(current)
                current = [line]
            elif current:
                current.append(line)
    if current:
        entries.append(current)
    return entries


def parse_name(entry_lines: list[str]) -> str | None:
    first = entry_lines[0]
    match = re.match(r"^(\d{3,4})\s*(.+)$", first)
    if not match:
        return None

    parts = [strip_address_tail(match.group(2))]
    if not re.match(r"^[A-Za-z(]", parts[0]):
        return None
    if parts[0].lower().startswith(("etc", "export/", "rmg", "n/a")):
        return None
    for line in entry_lines[1:6]:
        lowered = line.lower()
        if (
            "@" in line
            or re.search(r"\b\d{4,}\b", line)
            or any(token in lowered for token in ["road", "house", "plot", "sector", "block", "dhaka", "bangladesh", "gazipur", "savar", "avenue", "floor", "complex", "area", "c/a"])
            or "bd0" in lowered
            or "ra-" in lowered
        ):
            break
        parts.append(line)

    name = smart_title(" ".join(parts))
    name = re.sub(r"\s+", " ", name).strip()
    if not re.match(r"^[A-Za-z(]", name):
        return None
    if len(re.findall(r"[A-Za-z]", name)) < 5:
        return None
    return name if len(name) >= 4 else None


def parse_email(entry_text: str) -> str | None:
    match = re.search(r"([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})", entry_text, flags=re.I)
    return match.group(1).lower() if match else None


def parse_phone(entry_text: str) -> str | None:
    match = re.search(r"(\+?88)?0?1[0-9][0-9\- ]{7,}", entry_text)
    if not match:
        return None
    digits = re.sub(r"\D", "", match.group(0))
    return f"+{digits}" if digits else None


def detect_city(entry_text: str) -> str:
    for raw, city in sorted(BANGLADESH_CITIES.items(), key=lambda item: -len(item[0])):
        if re.search(rf"\b{re.escape(raw)}\b", entry_text, flags=re.I):
            return city
    return "Dhaka"


def infer_category_and_products(entry_text: str) -> tuple[str, list[str], str]:
    lowered = entry_text.lower()
    scores = {category: 0 for category in PRODUCT_TERMS}
    matched_terms: dict[str, list[str]] = {category: [] for category in PRODUCT_TERMS}

    for category, terms in PRODUCT_TERMS.items():
        for term in terms:
            if term in lowered:
                scores[category] += 1
                matched_terms[category].append(term)

    category = max(scores, key=lambda key: scores[key])
    if scores[category] == 0:
        category = "apparel"

    products = matched_terms[category][:3]
    if not products:
        default_map = {
            "apparel": ["t-shirts", "polo shirts", "woven garments"],
            "textiles": ["fabric", "yarn", "denim fabric"],
            "accessories": ["bags", "caps", "wallets"],
            "home": ["towels", "bedding", "home textiles"],
            "footwear": ["shoes", "casual footwear", "sneakers"],
        }
        products = default_map[category]

    subcategory = products[0]
    return category, products, subcategory


def build_real_suppliers(target_count: int) -> list[dict[str, Any]]:
    entries = extract_entries_from_pdf()
    suppliers: list[dict[str, Any]] = []
    seen_names: set[str] = set()

    for entry_lines in entries:
        entry_text = " ".join(entry_lines)
        if "bangladesh" not in entry_text.lower() and "dhaka" not in entry_text.lower():
            continue
        if "bd0" not in entry_text.lower():
            continue

        name = parse_name(entry_lines)
        if not name:
            continue
        normalized_name = re.sub(r"[^a-z0-9]+", "", name.lower())
        if normalized_name in seen_names:
            continue
        seen_names.add(normalized_name)

        category, products, subcategory = infer_category_and_products(entry_text)
        ordinal_match = re.match(r"^(\d{3,4})", entry_lines[0])
        ordinal = ordinal_match.group(1) if ordinal_match else str(len(suppliers) + 1)
        city = detect_city(entry_text)
        seed = f"{REAL_BATCH}:{ordinal}:{name}"
        price_low, price_high = CATEGORY_PRICE.get(category, (3.0, 8.0))
        moq_low, moq_high, moq_step = CATEGORY_MOQ.get(category, (300, 2000, 50))
        moq = int(round(numeric(f"{seed}:moq", moq_low, moq_high) / moq_step) * moq_step)
        risk_score = int(numeric(f"{seed}:risk", 22, 58))
        rating = numeric(f"{seed}:rating", 3.8, 4.6, 2)
        supplier = {
            "id": uuid_from_seed(seed),
            "name": name,
            "country": "Bangladesh",
            "city": city,
            "region": "South Asia",
            "category": category,
            "products": products,
            "description": f"{name} is listed in the Bangladesh Trade Portal textile exporters directory and is associated with {', '.join(products)} export activity.",
            "moq": moq,
            "lead_time_days": int(numeric(f"{seed}:lead", 18, 48)),
            "monthly_capacity": int(numeric(f"{seed}:capacity", 25000, 180000)),
            "unit_price_usd": numeric(f"{seed}:price", price_low, price_high, 2),
            "rating": rating,
            "risk_level": "low" if risk_score <= 30 else "medium" if risk_score <= 60 else "high",
            "risk_score": risk_score,
            "risk_notes": f"Public directory row. Operational fields are estimated for demo ranking and should be verified with the supplier before procurement.",
            "bgmea_certified": False,
            "certifications": [],
            "payment_terms": "To be confirmed with supplier",
            "contact_name": "Export Desk",
            "email": parse_email(entry_text),
            "phone": parse_phone(entry_text),
            "website": None,
            "notes": "Publicly sourced from Bangladesh Trade Portal textile exporters list. Name, location, and product direction are source-based; commercial fields are estimated for the demo.",
            "metadata": {
                "source_kind": "public_exporter_directory",
                "source_url": PDF_URL,
                "source_document": "Bangladesh Trade Portal textile exporters list",
                "source_entry_number": ordinal,
                "subcategory": subcategory,
                "raw_excerpt": entry_text[:500],
                "estimated_operational_fields": True,
                "import_batch": REAL_BATCH,
            },
        }
        suppliers.append(supplier)
        if len(suppliers) >= target_count:
            break

    return suppliers


def make_export_row(supplier: dict[str, Any], source_type: str, source_url: str | None, verified_at: str) -> dict[str, Any]:
    metadata = supplier.get("metadata", {})
    subcategory = metadata.get("subcategory")
    if not subcategory:
        subcategory = supplier.get("products", ["supplier products"])[0]

    risk_score = int(supplier["risk_score"])
    on_time_rate = max(72, min(98, int(98 - risk_score * 0.28)))
    quality_rating = float(supplier["rating"])

    return {
        "supplier_id": supplier["id"],
        "name": supplier["name"],
        "country": supplier["country"],
        "city": supplier["city"],
        "region": supplier["region"],
        "category": supplier["category"],
        "subcategory": subcategory,
        "products": ", ".join(supplier["products"]),
        "description": supplier["description"],
        "unit_price_usd": supplier["unit_price_usd"],
        "moq": supplier["moq"],
        "lead_time_days": supplier["lead_time_days"],
        "monthly_capacity": supplier["monthly_capacity"],
        "on_time_rate": on_time_rate,
        "quality_rating": quality_rating,
        "risk_score": risk_score,
        "risk_level": supplier["risk_level"],
        "bgmea_certified": supplier["bgmea_certified"],
        "certifications": ", ".join(supplier["certifications"]),
        "payment_terms": supplier["payment_terms"],
        "contact_name": supplier["contact_name"],
        "email": supplier["email"],
        "phone": supplier["phone"],
        "website": supplier["website"],
        "source_type": source_type,
        "source_url": source_url,
        "verified_at": verified_at,
        "embedding_status": "pending_local_hash_or_openai_embed",
        "metadata_json": json.dumps(supplier["metadata"], ensure_ascii=True),
        "notes": supplier["notes"],
    }


def autosize_sheet(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for index, value in enumerate(row, start=1):
            width = len(str(value)) if value is not None else 0
            widths[index] = min(max(widths.get(index, 0), width), 60)
    for index, width in widths.items():
        ws.column_dimensions[chr(64 + index) if index <= 26 else "A"].width = width + 2


def write_workbook(path: Path, title: str, rows: list[dict[str, Any]], sheet_name: str) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = title
    ws_summary["A1"].font = Font(size=14, bold=True)
    ws_summary["A3"] = "Total suppliers"
    ws_summary["B3"] = len(rows)
    ws_summary["D3"] = "Category"
    ws_summary["E3"] = "Supplier Count"
    ws_summary["G3"] = "Region"
    ws_summary["H3"] = "Supplier Count"

    category_counts = Counter(row["category"] for row in rows)
    region_counts = Counter(row["region"] for row in rows)
    country_counts = Counter(row["country"] for row in rows)

    ws_summary["A4"] = "Countries"
    ws_summary["B4"] = len(country_counts)
    ws_summary["A6"] = "Country"
    ws_summary["B6"] = "Supplier Count"

    for offset, (country, count) in enumerate(sorted(country_counts.items()), start=7):
        ws_summary[f"A{offset}"] = country
        ws_summary[f"B{offset}"] = count

    for offset, (category, count) in enumerate(sorted(category_counts.items()), start=4):
        ws_summary[f"D{offset}"] = category
        ws_summary[f"E{offset}"] = count

    for offset, (region, count) in enumerate(sorted(region_counts.items()), start=4):
        ws_summary[f"G{offset}"] = region
        ws_summary[f"H{offset}"] = count

    ws_rows = wb.create_sheet(sheet_name)
    ws_rows.append(EXPORT_COLUMNS)
    for row in rows:
        ws_rows.append([row[column] for column in EXPORT_COLUMNS])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    import csv

    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=EXPORT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def supplier_to_db_row(supplier: dict[str, Any]) -> dict[str, Any]:
    return {column: supplier.get(column) for column in DB_COLUMNS}


def upsert_supabase(rows: list[dict[str, Any]]) -> None:
    env = load_env_file()
    url = env.get("NEXT_PUBLIC_SUPABASE_URL") or env.get("SUPABASE_URL")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        raise RuntimeError("Missing SUPABASE URL or SUPABASE_SERVICE_ROLE_KEY in .env.local")

    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=minimal",
    }
    endpoint = f"{url}/rest/v1/suppliers?on_conflict=id"
    existing_response = requests.get(
        f"{url}/rest/v1/suppliers?select=id,name&limit=5000",
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
        },
        timeout=60,
    )
    if existing_response.status_code >= 300:
        raise RuntimeError(f"Could not load existing supplier names: {existing_response.status_code} {existing_response.text[:400]}")
    existing_rows = existing_response.json()
    existing_names = {str(row["name"]).strip().lower(): str(row["id"]) for row in existing_rows if row.get("name") and row.get("id")}

    unique_rows: list[dict[str, Any]] = []
    seen_names: set[str] = set()
    for row in rows:
        name_key = str(row["name"]).strip().lower()
        if not name_key or name_key in seen_names:
            continue
        existing_id = existing_names.get(name_key)
        if existing_id and existing_id != row["id"]:
            continue
        seen_names.add(name_key)
        unique_rows.append(row)

    db_rows = [supplier_to_db_row(row) for row in unique_rows]

    for index in range(0, len(db_rows), 100):
        batch = db_rows[index : index + 100]
        response = requests.post(endpoint, headers=headers, json=batch, timeout=120)
        if response.status_code >= 300:
            raise RuntimeError(f"Supabase upsert failed ({response.status_code}): {response.text[:800]}")


def count_supabase_rows() -> int | None:
    env = load_env_file()
    url = env.get("NEXT_PUBLIC_SUPABASE_URL") or env.get("SUPABASE_URL")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        return None

    response = requests.get(
        f"{url}/rest/v1/suppliers?select=id&limit=1",
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Prefer": "count=exact",
        },
        timeout=60,
    )
    if response.status_code >= 300:
        return None
    content_range = response.headers.get("Content-Range", "")
    match = re.search(r"/(\d+)$", content_range)
    return int(match.group(1)) if match else None


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate and optionally upsert expanded Sourcery supplier datasets.")
    parser.add_argument("--synthetic-count", type=int, default=500)
    parser.add_argument("--real-count", type=int, default=200)
    parser.add_argument("--upsert", action="store_true")
    args = parser.parse_args()

    synthetic = build_synthetic_suppliers(args.synthetic_count)
    real = build_real_suppliers(args.real_count)
    today = date.today().isoformat()

    synthetic_export = [make_export_row(row, "synthetic", None, today) for row in synthetic]
    real_export = [make_export_row(row, "public_web", PDF_URL, today) for row in real]

    write_workbook(SYNTHETIC_XLSX, "Sourcery Synthetic Supplier Dataset — Additional 500", synthetic_export, "Suppliers")
    write_workbook(REAL_XLSX, "Sourcery Public Supplier Dataset — 200 Real Rows", real_export, "Real Supplier Sample")

    combined = {
        "generated_on": today,
        "synthetic_count": len(synthetic),
        "real_count": len(real),
        "synthetic": synthetic,
        "real_public": real,
    }
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    COMBINED_JSON.write_text(json.dumps(combined, indent=2, ensure_ascii=True), encoding="utf-8")
    write_csv(COMBINED_CSV, synthetic_export + real_export)

    print(f"Generated {len(synthetic)} synthetic suppliers -> {SYNTHETIC_XLSX}")
    print(f"Generated {len(real)} public-sourced suppliers -> {REAL_XLSX}")
    print(f"Wrote combined JSON -> {COMBINED_JSON}")
    print(f"Wrote combined CSV -> {COMBINED_CSV}")

    if args.upsert:
        before = count_supabase_rows()
        upsert_supabase(synthetic + real)
        after = count_supabase_rows()
        print(f"Upserted {len(synthetic) + len(real)} suppliers into Supabase.")
        print(f"Live supplier count: before={before}, after={after}")


if __name__ == "__main__":
    main()
